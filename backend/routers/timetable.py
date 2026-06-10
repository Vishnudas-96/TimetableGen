from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from backend.database import get_connection
from backend.generator import TimetableGenerator
import io

router = APIRouter()

DAY_NAMES = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday"}

# ── Generation ─────────────────────────────────────────────────

@router.post("/generate")
def generate_timetable(data: dict = {}):
    section = data.get("section", "Middle Section")
    time_limit = data.get("time_limit", 120)
    gen = TimetableGenerator()
    result = gen.generate(section_name=section, time_limit_seconds=time_limit)
    gen.close()
    return result

@router.get("/logs")
def generation_logs():
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM generation_logs ORDER BY id DESC LIMIT 20"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── View endpoints ─────────────────────────────────────────────

@router.get("/class/{class_id}")
def class_timetable(class_id: int):
    gen = TimetableGenerator()
    result = gen.get_class_timetable(class_id)
    gen.close()
    return result

@router.get("/teacher/{teacher_id}")
def teacher_timetable(teacher_id: int):
    gen = TimetableGenerator()
    result = gen.get_teacher_timetable(teacher_id)
    gen.close()
    return result

@router.get("/master")
def master_timetable(section_id: int = None):
    """All classes timetable data in one call."""
    conn = get_connection()
    where = ""
    params = []
    if section_id:
        where = "WHERE c.section_id = ?"
        params.append(section_id)

    classes = conn.execute(
        f"SELECT * FROM classes {where} ORDER BY grade, division", params
    ).fetchall()

    result = []
    gen = TimetableGenerator()
    for cls in classes:
        tt = gen.get_class_timetable(cls["id"])
        result.append(tt)
    gen.close()
    conn.close()
    return result

# ── PDF Export ─────────────────────────────────────────────────

@router.get("/export/class/{class_id}/pdf")
def export_class_pdf(class_id: int):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    gen = TimetableGenerator()
    data = gen.get_class_timetable(class_id)
    gen.close()

    cls_info = data["class_info"]
    entries = data["entries"]
    slots = data["time_slots"]

    if not cls_info:
        return {"error": "Class not found"}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=16, fontName="Helvetica-Bold",
                                  alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle("sub", fontSize=10, fontName="Helvetica",
                                alignment=TA_CENTER, textColor=colors.grey, spaceAfter=12)

    # Build grid
    teaching_slots = [s for s in slots if not s["is_break"]]
    periods = sorted(set(s["period_number"] for s in teaching_slots))

    # Header row
    header = ["Day \\ Period"] + [f"P{p}" for p in periods]

    # Entry lookup: day_period -> entry
    entry_map = {(e["day_of_week"], e["period_number"]): e for e in entries}

    table_data = [header]
    for day in range(1, 6):
        row = [DAY_NAMES[day]]
        for p in periods:
            slot = next((s for s in slots if s["day_of_week"] == day and s["period_number"] == p), None)
            if slot and slot["is_break"]:
                row.append(slot["break_name"] or "Break")
            else:
                e = entry_map.get((day, p))
                if e:
                    row.append(f"{e['subject_short']}\n{e['teacher_short']}")
                else:
                    row.append("—")
        table_data.append(row)

    col_widths = [30*mm] + [22*mm] * len(periods)

    table = Table(table_data, colWidths=col_widths, rowHeights=20*mm)
    table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#6366f1")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",    (0, 1), (0, -1),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9ff")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
    ]))

    title = f"Timetable — {cls_info['grade']} Division {cls_info['division']}"
    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Generated by School Timetable Generator", sub_style),
        table,
    ]

    doc.build(elements)
    buffer.seek(0)

    filename = f"timetable_{cls_info['grade']}_{cls_info['division']}.pdf".replace(" ", "_")
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


# ── Excel Export ───────────────────────────────────────────────

@router.get("/export/class/{class_id}/excel")
def export_class_excel(class_id: int):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    gen = TimetableGenerator()
    data = gen.get_class_timetable(class_id)
    gen.close()

    cls_info = data["class_info"]
    entries = data["entries"]
    slots = data["time_slots"]

    if not cls_info:
        return {"error": "Class not found"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cls_info['grade']} {cls_info['division']}"

    teaching_slots = [s for s in slots if not s["is_break"]]
    periods = sorted(set(s["period_number"] for s in teaching_slots))

    # Styles
    header_fill = PatternFill("solid", fgColor="6366F1")
    day_fill    = PatternFill("solid", fgColor="EEF2FF")
    alt_fill    = PatternFill("solid", fgColor="F8F9FF")
    break_fill  = PatternFill("solid", fgColor="FEF3C7")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(periods)+1)}1")
    ws["A1"] = f"Timetable — {cls_info['grade']} Division {cls_info['division']}"
    ws["A1"].font = Font(bold=True, size=13, color="1E293B")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (periods)
    ws["A2"] = "Day \\ Period"
    ws["A2"].fill = header_fill
    ws["A2"].font = Font(bold=True, color="FFFFFF", size=10)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = thin_border

    for i, p in enumerate(periods, start=2):
        slot = next((s for s in teaching_slots if s["period_number"] == p), None)
        cell = ws.cell(row=2, column=i)
        cell.value = f"P{p}\n{slot['start_time']}-{slot['end_time']}" if slot else f"P{p}"
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.column_dimensions["A"].width = 14
    ws.row_dimensions[2].height = 30

    entry_map = {(e["day_of_week"], e["period_number"]): e for e in entries}

    for row_idx, day in enumerate(range(1, 6), start=3):
        ws.row_dimensions[row_idx].height = 36
        day_cell = ws.cell(row=row_idx, column=1, value=DAY_NAMES[day])
        day_cell.fill = day_fill
        day_cell.font = Font(bold=True, size=10)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.border = thin_border

        for col_idx, p in enumerate(periods, start=2):
            slot = next((s for s in slots if s["day_of_week"] == day and s["period_number"] == p), None)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if slot and slot["is_break"]:
                cell.value = slot["break_name"] or "Break"
                cell.fill = break_fill
                cell.font = Font(italic=True, color="92400E", size=8)
            else:
                e = entry_map.get((day, p))
                if e:
                    cell.value = f"{e['subject_short']}\n{e['teacher_short']}"
                    hex_color = e["subject_color"].lstrip("#") if e.get("subject_color") else "4A90E2"
                    cell.fill = PatternFill("solid", fgColor=hex_color + "33")
                    cell.font = Font(size=9, bold=True)
                else:
                    cell.value = "—"
                    cell.font = Font(color="CBD5E1", size=10)
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"timetable_{cls_info['grade']}_{cls_info['division']}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/all/excel")
def export_all_excel(section_id: int = None):
    """Export all classes in one Excel workbook, one sheet per class."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_connection()
    where = "WHERE section_id = ?" if section_id else ""
    params = [section_id] if section_id else []
    classes = conn.execute(
        f"SELECT * FROM classes {where} ORDER BY grade, division", params
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    gen = TimetableGenerator()
    for cls in classes:
        data = gen.get_class_timetable(cls["id"])
        entries = data["entries"]
        slots = data["time_slots"]

        ws = wb.create_sheet(title=f"{cls['grade']} {cls['division']}")
        teaching_slots = [s for s in slots if not s["is_break"]]
        periods = sorted(set(s["period_number"] for s in teaching_slots))
        entry_map = {(e["day_of_week"], e["period_number"]): e for e in entries}

        header_fill = PatternFill("solid", fgColor="6366F1")
        thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        ws["A1"] = f"{cls['grade']} {cls['division']}"
        ws["A1"].font = Font(bold=True, size=11)

        for i, p in enumerate(periods, start=2):
            c = ws.cell(row=2, column=i, value=f"P{p}")
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(horizontal="center")
            c.border = thin
            ws.column_dimensions[get_column_letter(i)].width = 15

        ws.column_dimensions["A"].width = 12

        for ri, day in enumerate(range(1, 6), start=3):
            ws.cell(row=ri, column=1, value=DAY_NAMES[day]).font = Font(bold=True)
            for ci, p in enumerate(periods, start=2):
                e = entry_map.get((day, p))
                cell = ws.cell(row=ri, column=ci)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if e:
                    cell.value = f"{e['subject_short']}\n{e['teacher_short']}"
                    cell.font = Font(size=8)
                else:
                    cell.value = "—"
            ws.row_dimensions[ri].height = 30

    gen.close()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=all_timetables.xlsx"}
    )


@router.get("/export/teacher/{teacher_id}/pdf")
def export_teacher_pdf(teacher_id: int):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    gen = TimetableGenerator()
    data = gen.get_teacher_timetable(teacher_id)
    gen.close()

    teacher = data["teacher_info"]
    entries = data["entries"]
    slots   = data["time_slots"]
    if not teacher:
        return {"error": "Teacher not found"}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)

    title_s = ParagraphStyle("t", fontSize=15, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    sub_s   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",      alignment=TA_CENTER,
                              textColor=colors.grey, spaceAfter=12)

    teaching_slots = [s for s in slots if not s["is_break"]]
    periods = sorted(set(s["period_number"] for s in teaching_slots))
    entry_map = {(e["day_of_week"], e["period_number"]): e for e in entries}

    header = ["Day"] + [f"P{p}" for p in periods]
    table_data = [header]
    for day in range(1, 6):
        row = [DAY_NAMES[day]]
        for p in periods:
            e = entry_map.get((day, p))
            row.append(f"{e['subject_short']}\n{e['grade']} {e['division']}" if e else "—")
        table_data.append(row)

    col_widths = [28*mm] + [22*mm]*len(periods)
    table = Table(table_data, colWidths=col_widths, rowHeights=18*mm)
    table.setStyle(TableStyle([
        ("BACKGROUND",      (0,0),(-1,0),  colors.HexColor("#10b981")),
        ("TEXTCOLOR",       (0,0),(-1,0),  colors.white),
        ("FONTNAME",        (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",        (0,0),(-1,0),  9),
        ("ALIGN",           (0,0),(-1,-1), "CENTER"),
        ("VALIGN",          (0,0),(-1,-1), "MIDDLE"),
        ("FONTNAME",        (0,1),(0,-1),  "Helvetica-Bold"),
        ("FONTSIZE",        (0,1),(-1,-1), 8),
        ("ROWBACKGROUNDS",  (0,1),(-1,-1), [colors.white, colors.HexColor("#f0fdf4")]),
        ("GRID",            (0,0),(-1,-1), 0.5, colors.HexColor("#e2e8f0")),
    ]))

    link_note = " [Link Teacher]" if teacher["is_link_teacher"] else ""
    elements = [
        Paragraph(f"Teacher Timetable — {teacher['name']}{link_note}", title_s),
        Paragraph(f"{teacher.get('subject_name','')} · {teacher.get('section_name','')} · "
                  f"{len(entries)} periods/week (max {teacher['max_periods_per_week']})", sub_s),
        table,
    ]
    doc.build(elements)
    buffer.seek(0)
    fname = f"timetable_teacher_{teacher['name'].replace(' ','_')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
